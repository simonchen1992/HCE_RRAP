import wx
import re
import gui
# html reading
from bs4 import BeautifulSoup
# xml reading
import xml.etree.ElementTree as ET
from os import walk,getcwd
import time
# pdf extraction libs
from pdfminer.pdfparser import  PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LTTextBoxHorizontal
from pdfminer.layout import LAParams
# output format: EXCEL
from openpyxl import Workbook
from openpyxl.styles import colors, PatternFill
# output format: PDF
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table,TableStyle, Frame, PageTemplate, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
# unique identifier
import random


ICSXML = {}
ICSTCMAPPING = {}
EXCEPTIONLIST = {}

"""
featureKey format: OPT_XXX_XXX_XXX; e.g. OPT_CDCVM_MOBILE_APPLICATION_DEVICE_PATTERN
featureValue format: true|false
resultKey format: T_XX_XX_XX; e.g. T_239_7_1_C02_01
resultValue format: pass|not applicable

ICSXML: {'ics_1.xml': {''OPT_XXX_XXX_XXX': 'False', 'OPT_XXX_XXX_XXX':....}, 'ics_2.xml': ......}
ICSTCMAPPING: {'ICS-1': [T_XX_XX_XX, T_XX_XX_XX, ...], 'ICS-2': ......}
"""


def footer(canvas, doc):
    canvas.saveState()
    pageNumber = ("%s" % canvas.getPageNumber())
    p = Paragraph('Page '+ pageNumber, normalStyle)
    w, h = p.wrap(1*inch, 1*inch)
    p.drawOn(canvas, 4*inch, 0.5*inch)
    p = Paragraph(otherInfo['Date'], normalStyle)
    w, h = p.wrap(2 * inch, 1 * inch)
    p.drawOn(canvas, 6.5 * inch, 0.5 * inch)
    p = Paragraph(otherInfo['Unique Identifier'], normalStyle)
    w, h = p.wrap(2 * inch, 1 * inch)
    p.drawOn(canvas, 1 * inch, 0.5 * inch)
    canvas.restoreState()


def header(canvas, doc):
    canvas.saveState()
    p = Paragraph("<font size=10 face='Courier'><b>RRAP VALIDATION REPORT - TOOL VERSION 1.0</b></font>", normalStyle)
    w, h = p.wrap(doc.width, doc.bottomMargin)
    p.drawOn(canvas, doc.leftMargin + doc.width - 6.2*inch, doc.topMargin + doc.height + 0.25*inch)
    #canvas.line(doc.leftMargin, doc.bottomMargin + doc.height + 0.3*inch, doc.leftMargin + doc.width, doc.bottomMargin + doc.height + 0.2*inch)
    canvas.restoreState()



class MianWindow(gui.MyFrame1):
    def searchCurrentFolder(self):
        path = getcwd()
        files = []
        for path, folder, file in walk(path):
            if re.search(r'^((?!/\.).)*$', path):  # Eliminate those temp folder and system folder, like .git
                for f in file:
                    files.append((f, path))
        return files

    def getAllIcsXml(self):
        for file, path in self.searchCurrentFolder():
            if re.search(r'ics_\d+.xml', file):
                tree = ET.parse(path + '/' + file)
                root = tree.getroot()
                if file not in ICSXML.keys():
                    ICSXML[file] = {}
                else:
                    self.executeLog.AppendText('%s: Repeated ics profile: %s. Please check if there is any collision in the folder.\n'
                                               % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), file))
                    raise Exception
                for element in root.findall('tagelement'):
                    ICSXML[file][element.find('Tag').text] = element.find('tagvalue').text

    def getIcsMapping(self):
        for file, path in self.searchCurrentFolder():
            if file == 'perso_ICS_Mapping.xml':
                tree = ET.parse(path + '/' + file)
                root = tree.getroot()
                for element in root.findall('ics'):
                    ICSTCMAPPING[element.attrib.get('profile')] = [tc.attrib.get('profile') for tc in element.findall('Test')]
            if file == 'filtering_exemption.xml':
                tree = ET.parse(path + '/' + file)
                root = tree.getroot()
                for element in root.findall('Rule'):
                    EXCEPTIONLIST[element.attrib.get('Description')] = [tc.text for tc in element.findall('TestCase')]

    def loadDataFromLayout(self, pattern, layout):
        output = []
        while pattern.search(layout) is not None:
            output.append(pattern.search(layout).group())
            layout = pattern.sub('', layout, 1)  # only remove once
        return output

    """
    POTENTIAL ISSUE:
    [featureKey, featureValue, resultKey, resultValue] are CASE SENSITIVE!!! e.g. 'true' != 'True'
    """
    #  load test tool results from pdf format
    def loadGalittReport(self, path):
        testToolResult = {}
        productFeatures = {}
        featureKey = []
        featureValue = []
        resultKey = []
        resultValue = []
        pValue = re.compile(r'true|false')  # Pattern for featureValue and resultValue
        pFeatureKey = re.compile(r'OPT_.*?(?=OPT_)|OPT_.*(?=)')  # Pattern for featureKey
        pResultKey = re.compile(r'T_\w+?_\w+?_\w+?_\d{2}')  # Pattern for resultKey
        otherInfo = {'Date':time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), 'Test Plan': '', 'Specification': '',
                     'Product': '', 'Vendor': '', 'Test Tool': 'Galitt'}
        with open(path, 'rb') as file:
            parser = PDFParser(file)
            doc = PDFDocument(parser)
            parser.set_document(doc)
            rsrcmgr = PDFResourceManager()
            laparams = LAParams()
            device = PDFPageAggregator(rsrcmgr, laparams=laparams)
            interpreter = PDFPageInterpreter(rsrcmgr, device)

            for page in PDFPage.create_pages(doc):
                interpreter.process_page(page)
                layout = device.get_result()
                # reset flag for page
                resultFlag = False
                featureFlag = False
                infoFlag = False
                try:
                    for text in layout:
                        if not isinstance(text, LTTextBoxHorizontal):
                            break
                        text = text.get_text().encode('utf-8').replace('\n', '')
                        # extract featureKEY when detects featureKey PATTERN
                        if text.find('OPT_') != -1:
                            featureFlag = True
                            featureKey.extend(self.loadDataFromLayout(pFeatureKey, text))
                        # extract resultKEY when detects resultKey PATTERN
                        if text.find('OPT_') == -1 and pResultKey.search(text) is not None:
                            resultFlag = True
                            resultKey.extend(self.loadDataFromLayout(pResultKey, text))

                    # If RRAP detects featureKEY in the page, RRAP will review this page again to get all featureVALUE
                    if featureFlag:
                        # extract device features VALUE: [true,false]
                        for text in layout:
                            if not isinstance(text, LTTextBoxHorizontal):
                                break
                            text = text.get_text().encode('utf-8').replace('\n', '')
                            featureValue.extend(self.loadDataFromLayout(pValue, text))
                            # # format featureValue
                            # for v in featureValue:
                            #     if v not in productFeatures['selection']:
                            #         raise Exception('2222')
                        # make sure featureKEY and featureVALUE share the same length
                        if len(featureKey) == len(featureValue):
                            # put featureKEY and featureValue in a dictionary and now finished for feature data extraction!
                            for i in range(len(featureKey)):
                                productFeatures[featureKey[i]] = featureValue[i]
                        else:
                            self.executeLog.AppendText('%s: The RRAP Tool collects %s ICS Questiones, but only with %s answer! Please check '
                                                    'if the chosen report and Test Tool vendor are correct. Or, there'
                                                    ' may have unexpected format for ICS Questiones.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                                                        len(featureKey), len(featureValue),))
                            raise Exception
                    # If RRAP detects resultKey in the page, RRAP will review this page again to get all resultValue
                    if resultFlag:
                        # extract test tool result VALUE: [true,false]
                        for text in layout:
                            if not isinstance(text, LTTextBoxHorizontal):
                                break
                            text = text.get_text().encode('utf-8').replace('\n', '')
                            resultValue.extend(self.loadDataFromLayout(pValue, text))
                            # format resultValue
                            for i, v in enumerate(resultValue):
                                if v not in ['pass', 'not applicable']:
                                    # if v in testToolResult['selection']:
                                    resultValue[i] = 'pass' if v == 'true' else 'not applicable'
                                    # else:
                                    #     raise Exception('222')
                        # make sure resultKey and resultValue share the same length
                        if len(resultKey) == len(resultValue):
                            # put resultKey and resultValue in a dictionary and now finished for test result data extraction!
                            for i in range(len(resultKey)):
                                testToolResult[resultKey[i]] = resultValue[i]
                        else:
                            self.executeLog.AppendText('%s: The test tool results keyword and results are not corresponding! Please check '
                                                       'if the chosen report and Test Tool vendor are correct. Or, there'
                                                       ' may have unexpected result in the test report.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                            raise Exception
                except AttributeError as e:
                    self.executeLog.AppendText('%s: Could not seek the keyword! Please check if the chosen report and '
                                               'Test Tool vendor are correct.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                    raise Exception
        # check if featureKEY is in ICS dictionary
        for f in featureKey:
            if f not in ICSXML['ics_1.xml'].keys():
                self.executeLog.AppendText('%s: Unexpected ICS QUESTION NAME: %s!\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), f))
                raise Exception
        # todo: check if resultKEY is in ICS dictionary

        return (productFeatures, testToolResult, otherInfo)


    def loadIccReport(self, path):
        testToolResult = {}  # 'selection' indicates original option for testToolResult.values()
        productFeatures = {}  # 'selection' indicates original option for productFeatures.values()
        featureKey = []
        featureValue = []
        resultKey = []
        resultValue = []
        pFeatureValue = re.compile(r'Yes|No(?!t)')  # Pattern for featureValue and resultValue
        pResultValue = re.compile(r'(Pass|Manual Pass|Fail|Manual Fail|Inconclusive|Not Applicable|Not Ran)(?!:)')  # Pattern for featureValue and resultValue
        pFeatureKey = re.compile(r'OPT.*?(?=:)')
        pResultKey = re.compile(r'T_\w+?_\w+?_\w+?_\d{2}')  # Pattern for resultKey
        otherInfo = {'Date':time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),'Test Plan':'', 'Specification':'',
                     'Product':'', 'Vendor':'', 'Test Tool':'ICCSolutions'}
        with open(path, 'rb') as file:
            parser = PDFParser(file)
            doc = PDFDocument(parser)
            parser.set_document(doc)
            rsrcmgr = PDFResourceManager()
            laparams = LAParams()
            device = PDFPageAggregator(rsrcmgr, laparams=laparams)
            interpreter = PDFPageInterpreter(rsrcmgr, device)

            for page in PDFPage.create_pages(doc):
                interpreter.process_page(page)
                layout = device.get_result()
                # reset flag for page
                resultFlag = False
                featureFlag = False
                infoFlag = False
                for text in layout:
                    if not isinstance(text, LTTextBoxHorizontal):
                        break
                    text = text.get_text().encode('utf-8').replace('\n', '')
                    # extract other info
                    try:
                        if text.find('Test Plan Name') != -1:
                            otherInfo['Test Plan'] = re.search(r'(?<=Test Plan Name:).*?(?=Company)', text).group().strip()
                        if text.find('Spec Name') != -1:
                            otherInfo['Specification'] = re.search(r'(?<=Spec Name:).*?(?=Spec)', text).group().strip() +\
                            re.search(r'(?<=Spec Version:).*?(?=Product)', text).group()
                        if text.find('Product Name') != -1:
                            otherInfo['Product'] = re.search(r'(?<=Product Name:).*?(?=Product)', text).group().strip() +\
                            re.search(r'(?<=Product Version:).*?(?=Vendor)', text).group()
                        if text.find('Vendor Name') != -1:
                            otherInfo['Vendor'] = re.search(r'(?<=Vendor Name:).*?(?=Contact)', text).group().strip()
                    except AttributeError as e:
                        self.executeLog.AppendText("%s: Couldn't pick device info automatically.\n" % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                    # extract featureKEY when detects featureKey PATTERN
                    if text.find('OPT ') != -1:
                        text.replace(' ', '_')
                        featureFlag = True
                        featureKey.extend(self.loadDataFromLayout(pFeatureKey, text))
                        # format featureKey
                        featureKey = [k.strip().replace(' ', '_') for k in featureKey]
                    # extract resultKEY when detects resultKey PATTERN
                    if text.find('OPT ') == -1 and pResultKey.search(text) is not None:
                        resultFlag = True
                        resultKey.extend(self.loadDataFromLayout(pResultKey, text))
                try:
                    # If RRAP detects featureKEY in the page, RRAP will review this page again to get all featureVALUE
                    if featureFlag:
                        # extract device features VALUE: [true,false]
                        for text in layout:
                            if not isinstance(text, LTTextBoxHorizontal):
                                break
                            text = text.get_text().encode('utf-8').replace('\n', '')
                            featureValue.extend(self.loadDataFromLayout(pFeatureValue, text))
                            # format featureValue
                            for i, v in enumerate(featureValue):
                                if v not in ['true', 'false']:  # if featureValue has not been formatted
                                    #if v in productFeatures['selection']:
                                    featureValue[i] = 'true' if v == 'Yes' else 'false'
                                    # else:
                                    #     self.executeLog.AppendText('%s: Unexpected feature\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                                    #     raise Exception
                        # make sure featureKEY and featureVALUE share the same length
                        if len(featureKey) == len(featureValue):
                            # put featureKEY and featureValue in a dictionary and now finished for feature data extraction!
                            for i in range(len(featureKey)):
                                productFeatures[featureKey[i]] = featureValue[i]
                        else:
                            self.executeLog.AppendText(
                                '%s: The feature keyword and results are not corresponding! Please check '
                                'if the chosen report and Test Tool vendor are correct. Or, there'
                                ' may have unexpected result in the test report.\n' % (
                                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                            raise Exception

                    # If RRAP detects resultKey in the page, RRAP will review this page again to get all resultValue
                    if resultFlag:
                        # extract test tool result VALUE: [true,false]
                        for text in layout:
                            if not isinstance(text, LTTextBoxHorizontal):
                                break
                            text = text.get_text().encode('utf-8').replace('\n', '')
                            resultValue.extend(self.loadDataFromLayout(pResultValue, text))
                            # format resultValue
                            for i, v in enumerate(resultValue):
                                if v not in ['pass', 'not applicable']:
                                    #if v in testToolResult['selection']:
                                    resultValue[i] = 'not applicable' if v == 'Not Applicable' else 'pass'
                                    # else:
                                    #     self.executeLog.AppendText('%s: Unexpected test tool results\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                                    #     raise Exception
                        # make sure resultKey and resultValue share the same length
                        if len(resultKey) == len(resultValue):
                            # put resultKey and resultValue in a dictionary and now finished for test result data extraction!
                            for i in range(len(resultKey)):
                                testToolResult[resultKey[i]] = resultValue[i]
                        else:
                            self.executeLog.AppendText('%s: The test tool results keyword and results are not corresponding! Please check '
                                                       'if the chosen report and Test Tool vendor are correct. Or, there'
                                                       ' may have unexpected result in the test report.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                            raise Exception
                except AttributeError as e:
                    self.executeLog.AppendText('%s: Could not seek the keyword! Please check if the chosen report and '
                                               'Test Tool vendor are correct.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                    raise Exception
        # del testToolResult['selection']
        # del productFeatures['selection']
        # check if featureKEY is in ICS dictionary
        for f in featureKey:
            if f not in ICSXML['ics_1.xml'].keys():
                self.executeLog.AppendText('%s: Unexpected ICS QUESTION NAME!\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                raise Exception
        # todo: check if resultKEY is in ICS dictionary
        return (productFeatures, testToolResult, otherInfo)

    def loadUlReport(self, path):
        testToolResult = {}  # 'selection' indicates original option for testToolResult.values()
        productFeatures = {}  # 'selection' indicates original option for productFeatures.values()
        featureKey = []
        featureValue = []
        resultKey = []
        resultValue = []
        pResultValue = re.compile(r'(failedSymbol|notApplicableSymbol|passedSymbol|inconclusiveSymbol|notExecutedSymbol)')  # Pattern for featureValue and resultValue
        pResultKey = re.compile(r'TC\w+?_\w+?_\w+?_\d{2}')  # Pattern for resultKey
        otherInfo = {'Date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), 'Test Plan': '', 'Specification': '',
                     'Product': '', 'Vendor': '', 'Test Tool': 'UL'}
        with open(path, 'r') as htmlFile:
            htmlHandler = htmlFile.read()
        html = BeautifulSoup(htmlHandler, "html.parser")
        html.prettify()
        try:
            #  extract otherInfo
            otherTable = html.find('table', id='tableTitleDetails')
            otherInfo['Specification'] = otherTable.find('td', 'tableKey', text='Specification:').find_next_sibling().get_text()
            otherInfo['Test Plan'] = otherTable.find('td', 'tableKey', text='Test Plan:').find_next_sibling().get_text()
            otherInfo['Product'] = otherTable.find('td', 'tableKey', text='Product:').find_next_sibling().get_text()
            otherInfo['Vendor'] = otherTable.find('td', 'tableKey', text='Vendor:').find_next_sibling().get_text()
            #  extract featureKey and featureValue
            cursor = html.find('h3', text='Implementation Conformance Statement')
            featureTable = [cursor.find_next_sibling(), cursor.find_next_sibling().find_next_sibling()]
            for table in featureTable:
                featureKey.extend([element.get_text() for element in table.find_all('td', 'detailsKeyWide')
                                  if element.get_text() not in ['Consumer Device CVM:', 'Other Options:', '']])
                featureValue.extend([element.get_text().strip().strip('\r').strip('\n') for element in table.find_all('td', 'charset_support')])
            # format featureKey
            for i, v in enumerate(featureKey):
                if v == 'qVSDC Track 2 Equivalent Data Format':
                    v = 'QVSDC_T2ED_WITH_MSD_VERIFICATION_VALUE_SUPPORT'
                if v == 'Get Data For Transaction Verification Log':
                    v = 'GET_DATA_COMMAND_FOR_TVL'
                v = v.replace('- ', '').replace(' ', '_').upper()
                formated = False
                for icsProfileName in ICSXML['ics_1.xml'].keys():
                    if v in icsProfileName:
                        featureKey[i] = icsProfileName
                        formated = True
                        break
                if formated == False:
                    self.executeLog.AppendText(
                        '%s: Unexpected ICS QUESTION NAME: %s!\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), v))
                    raise Exception
            #  format featureValue
            #  TODO: this part need to check with UL or BCN
            for i, v in enumerate(featureValue):
                if re.search(r'with MSD Verification Value', v):
                    v = 'Yes'
                elif re.search(r'without MSD Verification Value', v):
                    v = 'No'
                if v in ['Yes', 'No']:  # avoid unexpected featureValue
                    featureValue[i] = 'true' if v == 'Yes' else 'false'
            #  make sure featureKey and featureValue share the same length
            if len(featureKey) == len(featureValue):
                # put featureKEY and featureValue in a dictionary and now finished for feature data extraction!
                for i in range(len(featureKey)):
                    productFeatures[featureKey[i]] = featureValue[i]
                #  extra format of featureValue
                if productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] == 'Only with Online PIN CVM Priority':
                    productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] = 'false'
                    productFeatures['OPT_ONLINE_PIN_PRIORITY_SUPPORTED'] = 'true'
                #  TODO: this part need to check with UL or BCN
                elif productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] == 'false':
                    productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] = 'false'
                    productFeatures['OPT_ONLINE_PIN_PRIORITY_SUPPORTED'] = 'false'
                else:
                    productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] = 'true'
                    productFeatures['OPT_ONLINE_PIN_PRIORITY_SUPPORTED'] = 'false'


            else:
                self.executeLog.AppendText(
                    '%s: The feature keyword and results are not corresponding! Please check '
                    'if the chosen report and Test Tool vendor are correct. Or, there'
                    ' may have unexpected result in the test report.\n' % (
                        time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                raise Exception
            #  extract resultKey and resultValue
            cursor = html.find_all('h1', text='Test Case Summary')[0]
            resultTable = cursor.find_next_sibling().find_next_sibling()
            resultKey = self.loadDataFromLayout(pResultKey, resultTable.get_text())
            resultValue = self.loadDataFromLayout(pResultValue, ','.join([ele['class'][0] for ele in resultTable.find_all('div')]))
            #  format resultKey
            resultKey = [r.replace('C', '_', 1) for r in resultKey]
            #  format resultValue
            for i, v in enumerate(resultValue):
                resultValue[i] = 'not applicable' if v == 'notApplicableSymbol' else 'pass'
            # make sure resultKey and resultValue share the same length
            if len(resultKey) == len(resultValue):
                # put resultKey and resultValue in a dictionary and now finished for test result data extraction!
                for i in range(len(resultKey)):
                    testToolResult[resultKey[i]] = resultValue[i]
            else:
                self.executeLog.AppendText(
                    '%s: The test tool results keyword and results are not corresponding! Please check '
                    'if the chosen report and Test Tool vendor are correct. Or, there'
                    ' may have unexpected result in the test report.\n' % (
                        time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
                raise Exception
        except AttributeError as e:
            self.executeLog.AppendText('%s: Could not seek the keyword! Please check if the chosen report and '
                                       'Test Tool vendor are correct.\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
            raise Exception
        # todo: check if resultKEY is in ICS dictionary
        return (productFeatures, testToolResult, otherInfo)

    def testCaseFilter(self, productFeature):
        output = {}
        for icsProfile in ICSTCMAPPING.keys():
            testProfile = ICSTCMAPPING[icsProfile]
            icsProfile = icsProfile.lower().replace('-', '_') + '.xml'
            icsProfile = ICSXML[icsProfile]
            verdict = 'pass'
            for feature in icsProfile.keys():
                try:
                    if icsProfile[feature].lower() == 'true' and productFeature[feature].lower() == 'false':
                        verdict = 'not applicable'
                except AttributeError as e:
                    self.executeLog.AppendText('%s: ICS Question %s cannot be found in test report.\n' % (
                                                   time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), feature))
                    raise Exception
                else:
                    continue
            for tc in testProfile:
                output[tc] = verdict
        for condition in EXCEPTIONLIST.keys():
            f = re.search(r'OPT_.*?(?==)', condition).group()
            v = condition.replace(f, '').lower()
            v = re.search(r'true|false', v).group()
            #  if device feature meet exception condition, eliminate all involved test cases
            if productFeature[f] == v:
                for tc in EXCEPTIONLIST[condition]:
                    output[tc] = 'not applicable'
        return output

    #  generate unique identifier
    def genUniqueId(self):
        ret = ""
        for i in range(8):
            num = random.randint(0, 9)
            # num = chr(random.randint(48,57))
            letter = chr(random.randint(97, 122))
            Letter = chr(random.randint(65, 90))
            s = str(random.choice([num, letter, Letter]))
            ret += s
        return ret

    def genCompareReport(self, format):
        global ICSXML, ICSTCMAPPING, EXCEPTIONLIST, otherInfo, normalStyle
        ICSXML = {}
        ICSTCMAPPING = {}
        EXCEPTIONLIST = {}
        self.getAllIcsXml()
        self.getIcsMapping()

        vendorName =  self.inputChoice.GetString(self.inputChoice.CurrentSelection)
        if vendorName == 'Galitt':
            feature, reportResult, otherInfo = self.loadGalittReport(self.inputPath.GetPath())
        elif vendorName == 'ICCSolutions':
            feature, reportResult, otherInfo = self.loadIccReport(self.inputPath.GetPath())
        elif vendorName == 'UL':
            feature, reportResult, otherInfo = self.loadUlReport(self.inputPath.GetPath())
        rrapResult = self.testCaseFilter(feature)
        #  initial output format
        if format == 'pdf':
            # create pdf component
            pdfContent = []
            resultTableContent = []
            diffTableContent = []
            featureTableContent = []
            otherTableContent = []
            stylesheet = getSampleStyleSheet()
            normalStyle = stylesheet['Normal']
            headStyle = stylesheet['Heading1']
            pdfContent.append(Paragraph('<para fontSize=20 align=center><font face="times"><b>RRAP VALIDATION REPORT</b></font></para>', headStyle))
            otherInfo['Unique Identifier'] = self.genUniqueId()
            for keys in otherInfo.keys():
                otherTableContent.append([keys + ': ', otherInfo[keys]])

            otherTableContent.sort()
            pdfContent.append(Spacer(6 * inch, 4 * inch))
            pdfContent.append(Table(otherTableContent))
            pdfContent.append(PageBreak())
            pdfContent.append(Paragraph('<para fontSize=18 align=left><font face="times"><b>1. Project ICS Questions</b></font></para>', headStyle))
            pdfContent.append(Spacer(6 * inch, 0.5 * inch))
            for keys in feature.keys():
                featureTableContent.append([keys + ': ', feature[keys]])
            featureTableContent.sort()
            pdfContent.append(Table(featureTableContent))
            pdfContent.append(PageBreak())
        if format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = 'RRAP REPORT'
            ws['B1'] = 'Test Tool Result'
            ws['C1'] = 'RRAP Result'
            ws['D1'] = 'Comparision Result'
            row = 2
        Count = []
        for tcReal in reportResult.keys():
            verdict = ''
            for tcGen in rrapResult.keys():
                if tcReal in tcGen or tcGen in tcReal:
                    verdict = 'Correct' if reportResult[tcReal] == rrapResult[tcGen] else 'Not Correct'
                    Count.append(verdict)
                    break
                else:
                    continue
            # Exception if test cases name in report is not found in RRAP report
            if verdict == '':
                self.executeLog.AppendText('%s: Test Case %s cannot be found in test report.\n' % (
                    time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), tcReal))
                raise Exception
            # todo: decide output format: excel, word
            if format == 'excel':
                ws['A' + str(row)] = tcReal
                ws['B' + str(row)] = reportResult[tcReal]
                ws['C' + str(row)] = rrapResult[tcGen]
                ws['D' + str(row)] = verdict
                row += 1
            if format == 'pdf':
                resultTableContent.append([tcReal, reportResult[tcReal], rrapResult[tcGen], verdict])
                if verdict == 'Not Correct':
                    diffTableContent.append([tcReal, reportResult[tcReal], rrapResult[tcGen], verdict])
        if format == 'excel':
            # if not found in folder
            wb.save('RRAP Report.xlsx')
        if format == 'pdf':
            resultTableContent.sort()
            diffTableContent.sort()
            resultTableContent.insert(0, ['Test Case', 'Test Tool Result', 'RRAP Result', 'Comparision Result'])
            table = Table(resultTableContent, repeatRows=1, colWidths= 1.8 * inch)
            table.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,colors.grey),
                                       ('BACKGROUND', (0, 0), (-1, 0), colors.green)]))
            pdfContent.append(Paragraph('<para fontSize=18 align=left><font face="times"><b>2. Testing Results Overview</b></font></para>', headStyle))
            pdfContent.append(Spacer(6 * inch, 0.5 * inch))
            pdfContent.append(table)
            pdfContent.append(PageBreak())
            pdfContent.append(Paragraph('<para fontSize=18 align=left><font face="times"><b>3. Discrepancy Between RRAP '
                                        'Result and Test Tool Result</b></font></para>', headStyle))
            pdfContent.append(Spacer(6 * inch, 0.5 * inch))
            if Count.count('Not Correct') == 0:
                pdfContent.append(Paragraph('<para fontSize=18 align=left><font face="times">There is no discrepancy '
                                            'between RRAP Result and Test Tool Result.</font></para>', headStyle))
            else:
                pdfContent.append(Paragraph('<para fontSize=12 align=left><font face="times">There are %s discrepancies found '
                                            'between RRAP Result and Test Tool Result.</font></para>' %Count.count('Not Correct'), headStyle))
                diffTableContent.insert(0, ['Test Case', 'Test Tool Result', 'RRAP Result', 'Comparision Result'])
                diffTable = Table(diffTableContent, repeatRows=1, colWidths=1.8 * inch)
                diffTable.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('BACKGROUND', (0, 0), (-1, 0), colors.green)]))
                pdfContent.append(diffTable)

            pdf = SimpleDocTemplate('ValidationReport_%s.pdf'% otherInfo['Unique Identifier'])
            frame = Frame(pdf.leftMargin, pdf.bottomMargin, pdf.width, pdf.height, id='normal')
            templateF = PageTemplate(id='First', frames=frame, onPageEnd=footer)
            templateL = PageTemplate(id='Later', frames=frame, onPage=header, onPageEnd=footer)
            pdf.addPageTemplates([templateF, templateL])
            pdf.build(pdfContent)
            self.executeLog.AppendText('%s: Output finished!\n'
                                       '------------------------------------------\n' %(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))

def main():
    # # print re.search(r'(t|)(\d+)',bb).group()
    # #getAllIcsXml()
    # getIcsMapping()
    # print ICSTCMAPPING
    # # print ICSXML
    app = wx.App()
    main_win = MianWindow(None)
    main_win.Show()
    app.MainLoop()

    # getAllIcsXml()
    # feature, result = loadGalittReport('Galitt.pdf')
    # feature, result = loadIccReport('ICC_VCPCS.pdf')
    # rrap = testCaseFilter(feature)
    # genCompareReport(result, rrap, 'excel')

    # # testCaseFilter(p)

    # global normalStyle,headStyle
    # pdfContent = []
    # resultTableContent = []
    # stylesheet = getSampleStyleSheet()
    # normalStyle = stylesheet['Normal']
    # headStyle = stylesheet['Heading1']
    # pdfContent.append(Paragraph('<para fontSize=15 align=left><font face="times"><b>RRAP REPORT</b></font></para>', headStyle))
    # pdf = SimpleDocTemplate('test.pdf')
    # frame = Frame(pdf.leftMargin, pdf.bottomMargin, pdf.width, pdf.height, id='normal')
    # templateF = PageTemplate(id='First', frames=frame)
    # templateL = PageTemplate(id='Later', frames=frame, onPage=header ,onPageEnd=footer)
    #
    # for i in range(50):
    #     resultTableContent.append([i,1,1,1])
    # table = Table(resultTableContent, repeatRows=1)
    # pdfContent.append(PageBreak())
    # pdfContent.append(table)
    # pdf.addPageTemplates([templateF, templateL])
    # pdf.build(pdfContent)

    #loadTestToolPDF('Galitt.pdf')


# def html(path):
#     testToolResult = {}  # 'selection' indicates original option for testToolResult.values()
#     productFeatures = {}  # 'selection' indicates original option for productFeatures.values()
#     featureKey = []
#     featureValue = []
#     resultKey = []
#     resultValue = []
#     pResultValue = re.compile(r'(failedSymbol|notApplicableSymbol|passedSymbol|inconclusiveSymbol|notExecutedSymbol)')  # Pattern for featureValue and resultValue
#     pResultKey = re.compile(r'TC\w+?_\w+?_\w+?_\d{2}')  # Pattern for resultKey
#     otherInfo = {'Date': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), 'Test Plan': '', 'Specification': '',
#                  'Product': '', 'Vendor': '', 'Test Tool': 'UL'}
#     with open ('main-report1546718962545.html' , 'r') as htmlFile:
#         htmlHandler = htmlFile.read()
#     html = BeautifulSoup(htmlHandler, "html.parser")
#     html.prettify()
#     try:
#         #  extract otherInfo
#         otherTable = html.find('table', id='tableTitleDetails')
#         otherInfo['Specification'] = otherTable.find('td', 'tableKey', text='Specification:').find_next_sibling().get_text()
#         otherInfo['Test Plan'] = otherTable.find('td', 'tableKey', text='Test Plan:').find_next_sibling().get_text()
#         otherInfo['Product'] = otherTable.find('td', 'tableKey', text='Product:').find_next_sibling().get_text()
#         otherInfo['Vendor'] = otherTable.find('td', 'tableKey', text='Vendor:').find_next_sibling().get_text()
#         #  extract featureKey and featureValue
#         cursor = html.find('h3', text='Implementation Conformance Statement')
#         featureTable = [cursor.find_next_sibling(), cursor.find_next_sibling().find_next_sibling()]
#         for table in featureTable:
#             featureKey.extend([element.get_text() for element in table.find_all('td', 'detailsKeyWide')])
#             featureValue.extend([element.get_text() for element in table.find_all('td', 'charset_support')])
#         featureKey.remove('Consumer Device CVM:')
#         featureKey.remove('Other Options:')
#         # format featureKey
#         for i,v in enumerate(featureKey):
#             if v == 'qVSDC Track 2 Equivalent Data Format':
#                 v = 'QVSDC_T2ED_WITH_MSD_VERIFICATION_VALUE_SUPPORT'
#             if v == 'Get Data For Transaction Verification Log':
#                 v = 'GET_DATA_COMMAND_FOR_TVL'
#             v = v.replace('- ','').replace(' ', '_').upper()
#             formated = False
#             for icsProfileName in ICSXML['ics_1.xml'].keys():
#                 if v in icsProfileName:
#                     featureKey[i] = icsProfileName
#                     formated = True
#                     break
#             if formated == False:
#                 self.executeLog.AppendText(
#                    '%s: Unexpected ICS QUESTION NAME!\n' % (time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
#                 raise Exception
#         #  format featureValue
#         #  TODO: this part need to check with UL or BCN
#         for i,v in enumerate(featureValue):
#             if re.search(r'with MSD Verification Value', v):
#                 v = 'Yes'
#             elif re.search(r'without MSD Verification Value', v):
#                 v = 'No'
#             featureValue[i] = 'true' if v == 'Yes' else 'false'
#         #  make sure featureKey and featureValue share the same length
#         if len(featureKey) == len(featureValue):
#             # put featureKEY and featureValue in a dictionary and now finished for feature data extraction!
#             for i in range(len(featureKey)):
#                 productFeatures[featureKey[i]] = featureValue[i]
#             #  extra format of featureValue
#             if productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] == 'Only with Online PIN CVM Priority':
#                 productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] = 'false'
#                 productFeatures['OPT_ONLINE_PIN_PRIORITY_SUPPORTED'] = 'true'
#             #  TODO: this part need to check with UL or BCN
#             # elif productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] == 'Only with Online PIN CVM Priority':
#             #     productFeatures['OPT_CDCVM_PRIORITY_SUPPORTED'] = 'false'
#             #     productFeatures['OPT_ONLINE_PIN_PRIORITY_SUPPORTED'] = 'true'
#
#         else:
#             self.executeLog.AppendText(
#                 '%s: The feature keyword and results are not corresponding! Please check '
#                 'if the chosen report and Test Tool vendor are correct. Or, there'
#                 ' may have unexpected result in the test report.\n' % (
#                     time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
#             raise Exception
#         #  extract resultKey and resultValue
#         cursor = html.find_all('h1', text='Test Case Summary')[0]
#         resultTable = cursor.find_next_sibling().find_next_sibling()
#         resultKey = loadDataFromLayout(pResultKey, resultTable.get_text())
#         resultValue = loadDataFromLayout(pResultValue, ','.join([ele['class'][0] for ele in resultTable.find_all('div')]))
#         print len(resultValue)
#         #  format resultKey
#         resultKey = [r.replace('C' , '_', 1) for r in resultKey]
#         #  format resultValue
#         for i,v in enumerate(resultValue):
#             resultValue[i] = 'not applicable' if v == 'notApplicableSymbol' else 'pass'
#         # make sure resultKey and resultValue share the same length
#         if len(resultKey) == len(resultValue):
#             # put resultKey and resultValue in a dictionary and now finished for test result data extraction!
#             for i in range(len(resultKey)):
#                 testToolResult[resultKey[i]] = resultValue[i]
#         else:
#             self.executeLog.AppendText(
#                 '%s: The test tool results keyword and results are not corresponding! Please check '
#                 'if the chosen report and Test Tool vendor are correct. Or, there'
#                 ' may have unexpected result in the test report.\n' % (
#                     time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())))
#             raise Exception
#     except AttributeError as e:
#        print e
#     # todo: check if resultKEY is in ICS dictionary
#     return (productFeatures, testToolResult, otherInfo)
#
# def loadDataFromLayout(pattern, layout):
#     output = []
#     while pattern.search(layout) is not None:
#         output.append(pattern.search(layout).group())
#         layout = pattern.sub('', layout, 1)
#     return output
#
# def getAllIcsXml():
#     for file, path in searchCurrentFolder():
#         if re.search(r'ics_\d+.xml', file):
#             tree = ET.parse(path + '/' + file)
#             root = tree.getroot()
#             if file not in ICSXML.keys():
#                 ICSXML[file] = {}
#             for element in root.findall('tagelement'):
#                 ICSXML[file][element.find('Tag').text] = element.find('tagvalue').text

if __name__ == '__main__':
    main()
    #getAllIcsXml()
    #html(1)
    #text = 'Vendor Name: 23Contact'
    #print re.search(r'(?<=Vendor Name:).*(?=(Contact|))', text).group().strip()
    # with open('main-report1546718962545.html') as f:
    #     print f.read()

